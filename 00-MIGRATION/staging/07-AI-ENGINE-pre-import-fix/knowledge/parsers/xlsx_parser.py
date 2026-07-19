"""
app/knowledge/parsers/xlsx_parser.py
────────────────────────────────────────────────────────────────
Microsoft Excel (.xlsx) document parser.

Extracts:
  - Sheet names as section headers
  - All non-empty cell values
  - Tables as pipe-separated rows

Skips: formulas (evaluates to value), charts, images.

Requires: openpyxl
"""

from __future__ import annotations

import io
import logging

from app.knowledge.parsers.base import BaseParser, ParseResult

log = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 1000   # prevent memory issues on huge sheets
MAX_COLS_PER_ROW   = 50


class XlsxParser(BaseParser):
    supported_extensions = [".xlsx", ".xls"]
    doc_type = "spreadsheet"

    def parse(self, file_bytes: bytes, filename: str = "") -> ParseResult:
        try:
            import openpyxl
        except ImportError:
            return ParseResult(
                success=False,
                error="openpyxl not installed. Run: pip install openpyxl",
            )

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes),
                read_only=True,
                data_only=True,   # get values not formulas
            )

            sections: list[str] = []
            total_rows = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text: list[str] = []
                row_count = 0

                for row in ws.iter_rows(
                    max_row=MAX_ROWS_PER_SHEET,
                    values_only=True,
                ):
                    # Convert all cells to strings, skip fully empty rows
                    cells = [
                        str(cell).strip()
                        for cell in row[:MAX_COLS_PER_ROW]
                        if cell is not None and str(cell).strip()
                    ]
                    if cells:
                        rows_text.append(" | ".join(cells))
                        row_count += 1

                if rows_text:
                    sections.append(f"## Sheet: {sheet_name}")
                    sections.extend(rows_text)
                    total_rows += row_count

            content = "\n".join(sections)

            if not content.strip():
                return ParseResult(
                    success=False,
                    error="XLSX file contains no extractable data",
                )

            title = filename.replace(".xlsx", "").replace(".xls", "")
            title = title.replace("_", " ").strip() or "Spreadsheet"

            return ParseResult(
                success=True,
                content=content,
                title=title,
                doc_type=self.doc_type,
                page_count=len(wb.sheetnames),
                word_count=len(content.split()),
                extra_info={
                    "sheet_count": len(wb.sheetnames),
                    "total_rows":  total_rows,
                },
            )

        except Exception as exc:
            log.error("XLSX parse failed: %s", exc)
            return ParseResult(
                success=False,
                error=f"XLSX parse error: {exc}",
            )


xlsx_parser = XlsxParser()
